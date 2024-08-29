import pandas as pd
import openpyxl
import os

# Main folder path containing subfolders with .xlsm files
main_folder_path = r'C:\Users\sy.papadopoulos\OneDrive - Alumil S.A\Desktop\Countries'

def find_time_spend_row(sheet, column='G'):
    """Find the row above 'Time spent'."""
    for row in range(5, sheet.max_row + 1):
        if sheet[f'{column}{row}'].value == 'Time spent':
            return row - 1  # Return one row above 'Time spent'
    return None

def deduplicate_columns(headers):
    """Ensure DataFrame column headers are unique."""
    new_headers = []
    seen = {}
    for header in headers:
        new_header = header
        count = seen.get(header, 0)
        if count > 0:
            new_header = f"{header}_{count}"
        new_headers.append(new_header)
        seen[header] = count + 1
    return new_headers

def process_files_in_folder(folder_path):
    """Process all .xlsm files within a given folder."""
    all_data = pd.DataFrame()
    for file in os.listdir(folder_path):
        if file.endswith('.xlsm'):
            file_path = os.path.join(folder_path, file)
            wb = openpyxl.load_workbook(file_path, read_only=True, keep_vba=True, data_only=True)
            
            if len(wb.worksheets) < 5:
                print(f"Skipping {file} because it does not have enough sheets.")
                continue

            try:
                sheet1 = wb.worksheets[1]
                data1 = pd.DataFrame(sheet1.iter_rows(values_only=True, min_row=1, max_row=13, min_col=1, max_col=2))
                headers = data1.iloc[0]
                data1 = data1[1:]
                data1.columns = headers
                data1_transposed = data1.T
                data1_transposed.columns = data1_transposed.iloc[0]
                data1_transposed = data1_transposed[1:]
                data1_transposed.columns = deduplicate_columns(data1_transposed.columns.tolist())

                sheet2 = wb.worksheets[3]
                time_spend_row = find_time_spend_row(sheet2)
                if time_spend_row is not None:
                    data2 = pd.DataFrame(sheet2.iter_rows(values_only=True, min_row=5, max_row=time_spend_row, min_col=1, max_col=11))
                    headers = data2.iloc[0]
                    data2 = data2[1:]
                    data2.columns = headers
                    data2.columns = deduplicate_columns(data2.columns.tolist())

                    merged_df = pd.concat([data2.reset_index(drop=True), data1_transposed.reset_index(drop=True)], axis=1)
                    # Calculate the sum of "Total hours per month" for the current file
                    total_hours_sum = merged_df['Total hours per month'].sum()

                    # Create a new column with the sum value
                    merged_df['Sum Total hours per month'] = total_hours_sum

                    # Ensure 'Total hours per month' is numeric, converting any non-numeric to NaN
                    merged_df['Total hours per month'] = pd.to_numeric(merged_df['Total hours per month'], errors='coerce')

                    # Calculate the ratio. Use np.where to handle 0 or NaN in 'Total hours per month'
                    import numpy as np
                    merged_df['Hours Percentage'] = np.where(
                        merged_df['Total hours per month'] > 0,
                        merged_df['Total hours per month'] / merged_df['Sum Total hours per month'],
                        0
                    )

                    # Convert the ratio to percentage format (e.g., 0.5 to 50%)
                    merged_df['Hours Percentage'] = (merged_df['Hours Percentage'] * 100).round(2).astype(str) + '%'

                    # Replace any NaN or inf values in 'Hours Percentage' with '0%'
                    merged_df['Hours Percentage'].replace([np.inf, np.nan], '0%', inplace=True)
                    all_data = pd.concat([all_data, merged_df], ignore_index=True)

            except Exception as e:
                print(f"Error processing {file}: {e}")
                continue

    # Before saving the all_data DataFrame, drop the specified columns if they exist
    columns_to_drop = ['Στήλη1', 'Στήλη2', 'Στήλη3', 'Στήλη4', 'None_1', 'Null']
    # Use list comprehension to filter out columns that actually exist in the DataFrame
    columns_to_drop = [col for col in columns_to_drop if col in all_data.columns]

    # Drop the columns from the DataFrame
    if columns_to_drop:
        all_data.drop(columns=columns_to_drop, inplace=True)

    # Save the final DataFrame to a new Excel file
    folder_name = os.path.basename(os.path.normpath(folder_path))
    output_file_name = f"{folder_name}_merged_data.xlsx"
    output_file_path = os.path.join(folder_path, output_file_name)
    with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
        all_data.to_excel(writer, index=False)
    print(f"Data from folder '{folder_name}' merged and saved successfully.")

# Process each subfolder within the main folder
for subfolder in os.listdir(main_folder_path):
    subfolder_path = os.path.join(main_folder_path, subfolder)
    if os.path.isdir(subfolder_path):
        process_files_in_folder(subfolder_path)
